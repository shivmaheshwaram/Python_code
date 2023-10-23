import os
import re
import time
import xlrd
import pyexcel as p
#import win32com.client
from xlutils.copy import copy
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook


class XlsxUtility:

    def __init__(self, file_path, file_name):
        try:
            path_file_name = os.path.join(file_path, file_name)
            self.wb = load_workbook(filename=path_file_name, read_only=True)
            self.file_name = file_name
        except Exception as ex:
            print(ex.args)
        
    def read_row_from_sheet(self, sheet_idx=1, row_idx=2):
        sheets = self.wb.sheetnames
        sheet = self.wb.get_sheet_by_name(sheets[sheet_idx])
        read_row = []
        for cell in sheet[row_idx]:
            read_row.append(cell.value)
        return read_row

    def read_all_rows_from_sheet(self, sheet_idx=1):
        sheets = self.wb.sheetnames
        sheet = self.wb.get_sheet_by_name(sheets[sheet_idx])
        read_all_rows = []
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
            read_row = []
            for cell in row:
                read_row.append(cell.value)
            read_all_rows.append(read_row)
        return read_all_rows

    def read_cell_from_sheet(self, sheet_idx=1, row_idx=2, col_index=1):
        sheets = self.wb.sheetnames
        sheet = self.wb.get_sheet_by_name(sheets[sheet_idx])
        return sheet.cell(row=row_idx, column=col_index).value

    def insert_row_in_a_sheet_after(self, sheet_idx=1, row_idx=3):
        #row_idx : row will be inserted befor this param
        sheets = self.wb.sheetnames
        sheet = self.wb.get_sheet_by_name(sheets[sheet_idx])
        sheet.insert_rows(row_idx)
        self.wb.save(self.file_name)
        
    def update_single_row(self, sheet_idx=1, row_idx=3, values=[]):
        index = 0
        sheets = self.wb.sheetnames
        sheet = self.wb.get_sheet_by_name(sheets[sheet_idx])
        for cell in sheet[row_idx]:
            cell.value = values[index]
            index = index + 1
            if index >= sheet.max_column:
                break
        self.wb.save(self.file_name)
        
    def create_new_sheet(self, sheet_name='NewSheet'):
        self.wb.create_sheet(sheet_name)
        self.wb.save(self.file_name)
        
    def get_sheet_size(self, sheet_idx=0):
        sheets = self.wb.sheetnames
        sheet = self.wb.get_sheet_by_name(sheets[sheet_idx])
        return sheet.max_row, sheet.max_column

    def update_cell_value(self, sheet_idx=1, rownum=1, colnum=1, value=''):
        sheets = self.wb.sheetnames
        sheet = self.wb.get_sheet_by_name(sheets[sheet_idx])
        sheet.cell(row=rownum, column=colnum).value = value
        self.wb.save(self.file_name)

    def get_number_sheets(self):
        res = len(self.wb.sheetnames)
        return res

    def extract_sheets(self, target_folder):
        sheet_names = []
        try:
            base, extension = os.path.splitext(self.file_name)
            for sheet in self.wb.worksheets:
                new_wb = Workbook()
                ws = new_wb.active
                row = 0
                for row_data in sheet.iter_rows():
                    col_num = 1
                    row = row + 1
                    for row_cell in row_data:
                        type_of_cell = type(row_cell)
                        if type_of_cell.__name__ == 'EmptyCell':
                            column_letter = get_column_letter(col_num)
                            ws[column_letter + str(row)].value = ''
                        else:
                            ws[row_cell.coordinate].value = row_cell.value
                        col_num = col_num + 1
                sheet_names.append(re.sub(r"[^a-zA-Z]", "", sheet.title.lower()))
                new_wb.save(target_folder +'\\' + base +'_'+ '{0}.xlsx'.format(sheet.title))
        except Exception as ex:
            print(ex.args)
        return sheet_names

    def close_wb(self):
        self.wb.close()

class XlsUtility:

    def __init__(self, file_path, file_name):
        path_file_name = os.path.join(file_path, file_name)
        self.wb = xlrd.open_workbook(path_file_name, on_demand=True)
        self.file_name = file_name

    def get_number_sheets(self):
        res = len(self.wb.sheet_names())
        return res

    def extract_sheets(self, target_folder):
        base, extension = os.path.splitext(self.file_name)
        sheet_names = []
        for sheet in self.wb.sheets():  # cycles through each sheet in each workbook
            newwb = copy(self.wb)  # makes a temp copy of that book
            newwb._Workbook__worksheets = [worksheet for worksheet in newwb._Workbook__worksheets if
                                           worksheet.name == sheet.name]
            update_sheet = base + '_' + sheet.name.replace(' ', '')
            sheet_name = sheet.name.replace(' ', '').lower()
            newwb.save(target_folder + "\\" + update_sheet + '.xls')
            sheet_names.append(re.sub(r"[^a-zA-Z]", "", sheet_name))
        return sheet_names

    def run_macro(self, filename_macro, sheet_name):
        try:
            if os.path.exists(filename_macro):
                xl = win32com.client.Dispatch('Excel.Application')
                time.sleep(5)
                xl.DisplayAlerts = False
                xl.Visible = False
                wkb = xl.Workbooks.Open(Filename=os.path.abspath(filename_macro), ReadOnly=1)
                time.sleep(5)
                wkb.Sheets(sheet_name).Select()
                wkb.Application.Run("FormatDataCert")
                wkb.Save()
                wkb.Close(True)
                xl.Application.Quit()
                del xl
        except Exception as e:
            print(e.args)

    def get_sheet_size(self):
        sheet = self.wb.sheet_by_index(0)  # cycles through each sheet in each workbook
        return sheet.nrows, sheet.ncols

    def convert_xls_xlsx(self, src, dest):
        try:
            p.save_book_as(file_name=src, dest_file_name=dest)
        except Exception as e:
            print(e.args)

    def close_wb(self):
        self.wb.release_resources()


if __name__ == '__main__':
    work_book = XlsxUtility('Data/OutpuFile.xlsx')
    got_values = work_book.read_row_from_sheet(0, 2)
    print('APN_SOURCEL: ', got_values[0])
    print('STD_STREET_NUMBER: ', got_values[3])
    print('STD_UNIT_NUMBER: ', got_values[4])
    print('STD_STREET_NAME: ', got_values[5])
    print('STD_STREET_SUFFIX: ', got_values[6])
    print('STD_STREET_DIRECTION: ', got_values[66])
    print('STD_CITY: ', got_values[7])
    print('STD_STATE: ', got_values[8])
    print('STD_ZIP_CODE: ', got_values[9])
    print('LEGAL_DESCRIPTION: ', got_values[10])
    print('OWNER_LAST_NAME: ', got_values[21])
    print('OWNER_FIRST_NAME: ', got_values[22])
    work_book.update_value(0, 5, 1, got_values[0])

    # work_book.insert_row_in_a_sheet_after(1, 3)
    # work_book.update_single_row(1, 3, got_values)
    # update_values = work_book.read_row_from_sheet(1, 3)
    # print(update_values)
    # work_book.insert_row_in_a_sheet_after(1, 3)
    # print(work_book.read_row_from_sheet(0, 2))
    print(work_book.get_sheet_size())
