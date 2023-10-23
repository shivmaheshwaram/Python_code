import os
import csv
import pandas as pd
import configparser
global root_dir
from fuzzywuzzy import fuzz
from acquisition.CommonUtil.Common_Util import get_logger

log = get_logger('Header Validator')


class header_Validator():
    def __init__(self, prev_ver_file, cur_ver_file):
        try:
            self.current_headers = self.header_of(cur_ver_file)
            self.previous_headers = self.header_of(prev_ver_file)
            self.cur_ver_file = cur_ver_file
            self.prev_ver_file = prev_ver_file
            self.root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            self.propertyfilelocation = self.root_dir + "\environmentconfig.properties"
            self.configprop = configparser.RawConfigParser()
            self.configprop.read(self.propertyfilelocation)
            self.sqlconnectiontype = self.configprop.get("DATABASE", "conn_type")
            self.sqlservername = self.configprop.get("DATABASE", "server")
            self.sqlupdateheaderdb = self.configprop.get("DATABASE", "icp_update_db")
            self.sqlicpuser = self.configprop.get("DATABASE", 'icp_user')
            self.sqlicppassword = self.configprop.get("DATABASE", 'icp_password')
            self.sqltable = self.configprop.get("SQL", 'header_update_table')
            self.dbcorrecteditionheader = []
            self.dbcurrenteditionheader = []
        except Exception as e:
            log.error(e.args)

    def header_of(self, file_name):
        try:
            if 'xlsx' in file_name:
                df = pd.read_excel(file_name)
                fields = [i for i in df.columns]
            if ('tab' in file_name) or ('csv' in file_name) or ('txt' in file_name):
                with open(file_name, 'r') as f:
                    if 'tab' in file_name:
                        csv_f = csv.reader(f, delimiter='\t')
                    else:
                        csv_f = csv.reader(f)
                    fields = next(csv_f, None)
            return fields
        except Exception as e:
            log.error(e.args)

    def validate(self, icp, fips, county, edition, filename):
        try:
            log.info(self.current_headers)
            for current_header in self.current_headers:
                if 'Unnamed:' in current_header:
                    continue
                correct_header = self.character_match(current_header)
                # print(correct_header)
                if current_header != correct_header:
                    if self.match_columns(current_header, correct_header):
                        self.replace_correct_header_in_current_file(correct_header, current_header)
                        self.dbcorrecteditionheader.append(correct_header)
                        self.dbcurrenteditionheader.append(current_header)
                    else:
                        log.info("Column values doesn't match and < 70 percent (%s and %s)" % (current_header, correct_header))
            if len(self.dbcorrecteditionheader) > 0 or len(self.dbcurrenteditionheader) > 0:
                log.info("Updating the File_headers table with Correct headers")
        except Exception as e:
            log.error(e.args)

    def character_match(self, current_header):
        try:
            current_rate = 0
            current_item = current_header
            for item in self.previous_headers:
                rate = fuzz.ratio(current_header, item)
                if current_rate < rate:
                    current_rate = rate
                    current_item = item
            return current_item
        except Exception as e:
            log.error(e.args)

    def replace_correct_header_in_current_file(self, correct_header, wrong_header):
        try:
            lines = None
            if 'xlsx' in self.cur_ver_file:
                from openpyxl import load_workbook
                wb = load_workbook(self.cur_ver_file)
                sheets = wb.sheetnames
                ws = wb[sheets[0]]
                ch = 'A'
                x = chr(ord(ch) + self.current_headers.index(wrong_header)) + '1'
                ws[x] = correct_header
                wb.save(self.cur_ver_file)
            else:
                with open(self.cur_ver_file, "r") as current_file:
                    lines = current_file.readlines()
                    # print(wrong_header, correct_header)
                    lines[0] = lines[0].replace(wrong_header, correct_header)
                with open(self.cur_ver_file, "w") as f:
                    f.writelines(lines)
        except Exception as e:
            log.error(e.args)

    def match_columns(self, current_header, correct_header):
        try:

            good = 0
            bad = 0
            max_row_count = 100
            with open(self.cur_ver_file, "r") as f:
                if 'xlsx' in self.cur_ver_file:
                    df = pd.read_excel(self.cur_ver_file)
                    column_list = [i for i in df[current_header]][:max_row_count]
                else:
                    if 'tab' in self.cur_ver_file:
                        csv_f = csv.DictReader(f, fieldnames=self.current_headers, delimiter='\t')
                    else:
                        csv_f = csv.DictReader(f, fieldnames=self.current_headers)

                    column_list = [row[current_header] for row in csv_f if current_header not in row[current_header]]
                with open(self.prev_ver_file, "r") as pf:
                    if 'xlsx' in self.cur_ver_file:
                        df = pd.read_excel(self.prev_ver_file)
                        column_list_p = [i for i in df[correct_header]][:max_row_count]
                    else:
                        if 'tab' in self.prev_ver_file:
                            csv_pf = csv.DictReader(pf, fieldnames=self.previous_headers, delimiter='\t')
                        else:
                            csv_pf = csv.DictReader(pf, fieldnames=self.previous_headers)
                        column_list_p = [row[correct_header] for row in csv_pf if correct_header not in row[correct_header]]
                    column_count = 0
                    for k, v in zip(column_list, column_list_p):
                        column_count = column_count + 1
                        if column_count > max_row_count:
                            break;
                        # print(k)
                        if fuzz.ratio(str(k), str(v)) > 50:
                            good += 1
                            # print("fuzz %d", k)
                        elif k in column_list_p:
                            good += 1
                            # print("search %s",k)
                        else:
                            bad += 1
                            # print("bad %s",k)
                    log.info("Header match between %s and %s" % (current_header, correct_header))
                    total_data = good+bad
                    log.info("Good_data: {} Bad_data: {} Total data for the column: {}" .format(good, bad, total_data))
                    if (float(good) / float(good + bad)) * 100.0 > 70.0:
                        log.info("Good:{}, Total Match is > 70% with prior column".format(good))
                        return True
                    else:
                        return False
        except Exception as e:
            log.error(e.args)

    def header_replace(self, state, county, prev, curr, folderPath, icp, fips, edition):
        try:
            # fileDestinationPath = '//10.208.200.247//edgqn1_dideg//QA Projects//Tax Roll//Data Prep//'
            fileDestinationPath = '//Vnx2566cifs3/EDGQN1_TXRACQ/Data Prep/'
            filedownloadPath = folderPath
            pe = 'Prior Edition'
            if prev and curr:
                prev_file = fileDestinationPath+"{0}//{1}//{2}//{3}".format(state, county, pe, prev)
                curr_file = filedownloadPath +curr
                hv = header_Validator(prev_file, curr_file)
                hv.validate(icp, fips, county, edition, curr)
        except Exception as e:
            log.error(e.args)
